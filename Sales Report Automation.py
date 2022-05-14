# -*- coding: utf-8 -*-
"""
Created on Sat Oct 16 22:59:18 2021

@author: Ryan
"""

from __future__ import print_function
import excel2img
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
import base64
import os.path
from google.oauth2.credentials import Credentials
import numpy as np
from datetime import datetime


# This call structure is necessary to install the google libraries, and is provided
# below given the distant connection between the library name called for installation vs import:
#
# pip install --upgrade google-api-python-client google-auth-httplib2 google-auth-oauthlib


def gather_input_info():
    '''
    Gather information from the user for customizing the program outputs.

    Returns
    -------
    image_date : String.
        String of the date for the end of the period. Used to label each image
        produced. Ex. 'NOV 15' for the November 1 - November 15 reporting period.
    sheetname : String.
        Name of the worksheet in the Excel workbook for the reporting period.
        Ex. 'NOV 1 to 15' for the November 1 - November 15 reporting period.
    period_start_date : String.
        The starting date of the reporting period for customizing the outgoing
        email. Ex. 'November 1st' for the November 1 - November 15 reporting period.
    period_end_date : String.
        The ending date of the reporting period for customizing the outgoing
        email. Ex. 'November 15th' for the November 1 - November 15 reporting period.
    current_full_report : DataFrame.
        The current reporting period transactions report in DataFrame format.
    vendor_info_df : DataFrame.
        The vendor information table with vendor ID, contact information, and
        payment details in DataFrame format.

    '''
    square_download_filename = input('Please paste the name of the downloaded square file with item transaction details: ')
    if square_download_filename[-3:] != 'csv':
        square_download_filename = square_download_filename + '.csv'
    image_date = input('Please type the date to mark the screenshots with: ')
    last_sheetname = input('Please type the name of the sheet from the last period: ')
    new_sheetname = input('Please type the name of the sheet for the current period being reported: ')
    period_start_date = input('Please type the date to mark the start of'
                              ' the period for the email blast:')
    period_end_date = input('Please type the date to mark the end of'
                            'the period for the email blast:')
    last_full_report = pd.read_excel(f'C:/Users/Ryan/BevCol/The Beverly Collective Sales {datetime.now().year}.xlsx',
                                     last_sheetname)
    vendor_info_df = pd.read_excel(f'C:/Users/Ryan/BevCol/The Beverly Collective Sales {datetime.now().year}.xlsx',
                                   'VENDOR LIST')
    try:
        former_vendor_index = vendor_info_df.index[vendor_info_df['NAME'] == 'FORMER VENDORS'][0]
        vendor_info_df = vendor_info_df[:former_vendor_index]
    except IndexError:
        pass
    all_vendors_list = [code for code in vendor_info_df['VENDOR CODE'] if code is not np.nan]
    return (image_date, new_sheetname, period_start_date,
            period_end_date, last_full_report, vendor_info_df, all_vendors_list, square_download_filename)


def create_image_attachments(current_full_report, image_date):
    '''
    Creates an image screenshoting the transactions of each vendor who made sale
    transactions during the current reporting period and saves the image to the local
    directory.

    Parameters
    ----------
    current_full_report : Pandas DataFrame
        DataFrame containing the current reporting period transactions report.
    image_date : str
        The final report date to include in the title of each saved image file.

    Returns
    -------
    None.

    '''
    for vendor in current_full_report['VENDOR ID'].unique():
        temp_df = current_full_report[current_full_report['VENDOR ID']==vendor]
        with pd.ExcelWriter(f'{vendor}_temp.xlsx') as writer:
            temp_df.to_excel(writer, sheet_name='Sheet1', index=False)
        # Auto-adjust columns' width
            for column in temp_df:
                column_width = max(temp_df[column].astype(str).map(len).max(), len(column))
                if column_width < 10:
                    column_width = 10
                col_idx = temp_df.columns.get_loc(column)
                writer.sheets['Sheet1'].set_column(col_idx, col_idx, column_width)
            writer.save()
            wb = load_workbook(filename = f'{vendor}_temp.xlsx')
            ws = wb.active
            for i in range(len(temp_df)):
                _cell = ws[f'I{i+2}']
                _cell.number_format = 'm/d/yyyy'
                _cell = ws[f'A{i+2}']
                _cell.alignment = Alignment(horizontal='center')
                _cell = ws[f'J{i+2}']
                _cell.number_format = 'h:mm AM/PM'
                for currency_col in ['E', 'F', 'G', 'H']:
                    _cell = ws[f'{currency_col}{i+2}']
                    _cell.number_format = '$#,##0.00_);[Red]($#,##0.00)'
            wb.save(f'{vendor}_temp.xlsx')
            # writer.close()
        if vendor == 'BEE':
            continue
            #os.rename(f'{vendor}_temp.xlsx', f'{vendor} Sales {image_date}.xlsx')
        else:
            excel2img.export_img(f'{vendor}_temp.xlsx',
                                 f'{vendor} Sales {image_date}.png',
                                 'Sheet1',
                                 None)
            os.remove(f'{vendor}_temp.xlsx')
    return


def create_message(sender, to, subject, message_text):
    """
    Create a message for an email.
    Code from: https://developers.google.com/gmail/api/guides/drafts#python

    Parameters
    -------
    sender : str
        Email address of the sender.
    to : str
        Email address of the receiver.
    subject : str
        The subject of the email message.
    message_text : str
        The text of the email message.

    Returns
    -------
    return_dict : dict
        A dictionary object containing a base64url encoded email object.

    """
    message = MIMEMultipart(message_text)
    message = MIMEText(message_text)
    message['to'] = to
    message['from'] = sender
    message['subject'] = subject
    return_dict =  {'raw': base64.urlsafe_b64encode(message.as_string().
                                            encode('UTF-8')).
                    decode('ascii')}
    return return_dict


def create_draft(service, user_id, subject, message_body, to):
    """
    Create and insert a draft email. Print the returned draft's message and id.
    Some code of this function sourced from:
    https://developers.google.com/gmail/api/guides/drafts

    Parameters
    -------
    service : API instance
        Authorized Gmail API service instance.
    user_id : str
        User's email address. The special value "me" can be used to indicate the authenticated user.
    message_body : MIME object
        The body of the email message, including headers.
    to : str
        Message recipient's email address

    Returns
    -------
    draft : MIME object
        Draft MIME object, including draft id and message meta data.

    """
    message = MIMEText(message_body)
    message['to'] = to
    message['from'] = user_id
    message['subject'] = subject
    draft_message = {'message': {'raw': base64.urlsafe_b64encode(message.as_string().encode('UTF-8')).decode('ascii')}}
    draft = service.users().drafts().create(userId=user_id, body=draft_message).execute()

    print(f'Draft id: {draft["id"]}')
    print(f'Draft message: {draft["message"]}')

    return draft



def create_email_text(vendor_firstname, date_start, date_end, amount, payment_method, email):
    """
    Create full text of an email tailored to the details of the vendor and reporting period.

    Parameters
    -------
    vendor_firstname : str
        First name of the vendor receiving the email.
    date_start : str
        String representation of the first date of the reporting period.
    date_end : str
        String representation of the last date of the reporting period.
    amount : float
        Amount earned by the vendor for the reporting period.
    payment_method : str
        The preferred payment method of the vendor.
    email : str
        Message recipient's email address.

    Returns
    -------
    text : str
        Full text of email populated with vendor and payment variables for a tailored message.

    """
    text = f'''Hi {vendor_firstname},\n
I have attached a screenshot of your sales at The Beverly Collective for the period of {date_start} through {date_end}.\n
I will make the payment of {"${:,.2f}".format(amount)} to you via {payment_method}.\n
Thank you, and I hope you are having a great week!\n
Best,
Monica
'''
    return text


def create_zero_sales_email_text(vendor_firstname, date_start, date_end, email):
    """
    Create full text of an email tailored to the details of the vendor and reporting period
    when the vendor made no sales for the reporting period.

    Parameters
    -------
    vendor_firstname : str
        First name of the vendor receiving the email.
    date_start : str
        String representation of the first date of the reporting period.
    date_end : str
        String representation of the last date of the reporting period.
    email : str
        Message recipient's email address.

    Returns
    -------
    text : str
        Full text of email populated with vendor and payment variables for a tailored message.

    """
    text = f'''Hi {vendor_firstname}!\n
I am sorry to say that you didn't have any sales at The Beverly Collective for the period of {date_start} through {date_end}.\n
Thank you, and I hope you are having a great week!\n
Best,
Monica'''
    return text


def calculate_summary_amount(current_full_report):
    '''
    Create an aggregate sales transactions report providing top-level sales
    information for each vendor.

    Parameters
    ----------
    current_full_report : DataFrame.
        DataFrame containing the current reporting period transactions report.

    Returns
    -------
    grouped_df : DataFrame.
        DataFrame containing a summary output of the aggregate net sales, sales tax,
        vendor earnings, and The Beverly Collective commission for each vendor.

    '''
    grouped_df = current_full_report.groupby(by='VENDOR ID').agg('sum')
    grouped_df = grouped_df.drop(columns='TRANSACTION')
    grouped_df.to_csv('C:/Users/Ryan/BevCol/Bev_Sales_Summary.csv')
    return grouped_df


def create_vendor_email_drafts(service, current_full_report, grouped_df, vendor_info_df, period_start_date, period_end_date, image_date, all_vendors_list):
    '''
    Creates an outgoing email draft in the Gmail account of the owner of The Beverly
    Collective containing customized subject and body. Some code of this function
    sourced from: https://developers.google.com/gmail/api/guides/drafts

    Parameters
    ----------
    service : Gmail API Connection
        An active connection with live certification and token to the Gmail API.
    current_full_report : Pandas DataFrame
        DataFrame containing the current reporting period transactions report.
    grouped_df : Pandas DataFrame
        DataFrame containing a summary output of the aggregate net sales, sales tax,
        vendor earnings, and The Beverly Collective commission for each vendor.
    vendor_info_df : Pandas DataFrame
        The vendor information table with vendor ID, contact information, and
        payment details in DataFrame format.
    period_start_date : str
        The starting date of the reporting period for customizing the outgoing
        email. Ex. 'November 1st' for the November 1 - November 15 reporting period.
    period_end_date : str
        The ending date of the reporting period for customizing the outgoing
        email. Ex. 'November 15th' for the November 1 - November 15 reporting period.
    image_date : str
        String of the date for the end of the period. Used to label each image
        produced. Ex. 'NOV 15' for the November 1 - November 15 reporting period.
    all_vendors_list : list
        List of all current vendors for whom communication is needed.

    Returns
    -------
    None.

    '''
    for vendor in all_vendors_list:
        print(f'Creating draft for {vendor}.')
        vendor_email = str(vendor_info_df[vendor_info_df['VENDOR CODE']==vendor]
                           ['EMAIL'].values)[2:-2]
        vendor_firstname = str(vendor_info_df[vendor_info_df['VENDOR CODE']==vendor]
                               ['NAME']).split()[1]
        if vendor in list(grouped_df.index):
            vendor_amount = grouped_df['VENDOR PAYOUT'][vendor]
            payment_method = str(vendor_info_df[vendor_info_df['VENDOR CODE']==vendor]
                                 ['PAYMENT'].values[0])
            create_draft(service=service,
                         user_id='monica.vercillo@gmail.com',
                         subject=f'{vendor} Update {image_date} The Beverly Collective',
                         message_body=create_email_text(vendor_firstname,
                                                        period_start_date,
                                                        period_end_date,
                                                        vendor_amount,
                                                        payment_method,
                                                        vendor_email),
                         to=vendor_email)
        else:
            create_draft(service=service,
                         user_id='monica.vercillo@gmail.com',
                         subject=f'{vendor} Update {image_date} The Beverly Collective',
                         message_body=create_zero_sales_email_text(vendor_firstname,
                                                                   period_start_date,
                                                                   period_end_date,
                                                                   vendor_email),
                         to=vendor_email)
    return


def establish_gmail_api_connection():
    '''
    Establish Gmail API Connection, prompting user for app permission and
    returning active API connection. Most code of this function sourced from:
    https://developers.google.com/gmail/api/quickstart/python

    Returns
    -------
    service : Gmail API Connection.
        An active connection with live certification and token to the Gmail API.

    '''
    if os.path.exists('C:/Users/Ryan/BevCol/token.json'):
        os.remove('C:/Users/Ryan/BevCol/token.json')
    # If modifying these scopes, delete the file token.json.
    SCOPES = ['https://www.googleapis.com/auth/gmail.readonly',
              'https://www.googleapis.com/auth/gmail.compose',
              'https://www.googleapis.com/auth/gmail.insert',
              'https://www.googleapis.com/auth/gmail.settings.sharing',
              'https://www.googleapis.com/auth/gmail.addons.current.message.metadata',
              'https://www.googleapis.com/auth/gmail.addons.current.action.compose',
              'https://www.googleapis.com/auth/gmail.addons.current.message.action',
              'https://www.googleapis.com/auth/gmail.labels']
    creds = None
    # The file token.json stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
    if os.path.exists('token.json'):
        creds = Credentials.from_authorized_user_file('token.json', SCOPES)
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        # Save the credentials for the next run
        with open('token.json', 'w') as token:
            token.write(creds.to_json())

    service = build('gmail', 'v1', credentials=creds)

    # Call the Gmail API
    results = service.users().labels().list(userId='me').execute()
    labels = results.get('labels', [])

    if not labels:
        print('No labels found.')
    else:
        print('Labels:')
        for label in labels:
            print(label['name'])
    return service


def preprocess_item_details_square_file(square_download_filename,
                                        first_transaction_num,
                                        vendor_info_df,
                                        new_sheetname):
    '''
    Clean and format item details transaction report produced by Square for the
    reporting period.

    Parameters
    ----------
    square_download_filename : str
        The file name of the item details transaction report downloaded from Square.
    first_transaction_num : int
        The unique transaction number of the first transaction of the period based on the last
        transaction number from the last reporting period.
    vendor_info_df : Pandas DataFrame
        The vendor information table with vendor ID, contact information, and
        payment details in DataFrame format.
    new_sheetname : str
        The name of the new sheet to create in the sales Excel workbook containing the
        latest reporting period transactions.

    Returns
    -------
    item_transactions_df : Pandas DataFrame
        A clean DataFrame with the information from the current reporting period as was inserted
        into the sales Excel workbook.

    '''
    item_transactions_df = pd.read_csv(os.path.join('C:/Users/Ryan/Downloads/', square_download_filename))
    item_transactions_df = item_transactions_df[['Category',
                                                 'Item',
                                                 'Qty',
                                                 'Net Sales',
                                                 'Tax',
                                                 'Date',
                                                 'Time']]
    item_transactions_df.rename(columns={'Category': 'VENDOR ID',
                                         'Item': 'ITEM CODE',
                                         'Net Sales': 'PRICE',
                                         'Tax': 'SALES TAX',
                                         'Date': 'DATE',
                                         'Time': 'TIME'},
                                inplace=True)
    item_transactions_df['ITEM DESCRIPTION'] = [' '.join(item.split(' ')[1:]) for item in item_transactions_df['ITEM CODE']]
    item_transactions_df['VENDOR ID'] = [code.upper() for code in item_transactions_df['VENDOR ID']]
    item_transactions_df['ITEM CODE'] = [item.split(' ')[0] for item in item_transactions_df['ITEM CODE']]
    item_transactions_df.sort_values(by=['DATE', 'TIME'], inplace=True)
    item_transactions_df.reset_index(drop=True, inplace=True)
    item_description_w_quantity_list = []
    transaction_num_list = []
    bev_payout_list = []
    vendor_payout_list = []
    for index, row in item_transactions_df.iterrows():
        if index == 0:
            current_transaction_num = first_transaction_num
            transaction_num_list.append(current_transaction_num)
        else:
            if row['TIME']==item_transactions_df.at[index-1, 'TIME'] and row['DATE']==item_transactions_df.at[index-1, 'DATE']:
                transaction_num_list.append(current_transaction_num)
            else:
                current_transaction_num += 1
                transaction_num_list.append(current_transaction_num)
        if row['Qty'] == 1:
            item_description_w_quantity_list.append(row['ITEM DESCRIPTION'])
        else:
            item_description_w_quantity_list.append(row['ITEM DESCRIPTION'] + f' x{str(int(row["Qty"]))}')
        try:
            bev_payout_val = round(np.array(vendor_info_df[vendor_info_df['VENDOR CODE']==row['VENDOR ID']]['COMMISSION'])[0] * float(row['PRICE'][1:]), 2)
            bev_payout_list.append(bev_payout_val)
            vendor_payout_val = round(float(row['PRICE'][1:])-bev_payout_val, 2)
            vendor_payout_list.append(vendor_payout_val)
        except IndexError:
            bev_payout_list.append('Error')
            vendor_payout_list.append('Error')
    item_transactions_df['BEV PAYOUT'] = bev_payout_list
    item_transactions_df['VENDOR PAYOUT'] = vendor_payout_list
    item_transactions_df['TRANSACTION'] = transaction_num_list
    item_transactions_df['ITEM DESCRIPTION'] = item_description_w_quantity_list
    item_transactions_df.drop(columns=['Qty'], inplace=True)
    item_transactions_df = item_transactions_df[['TRANSACTION',
                                                 'VENDOR ID',
                                                 'ITEM CODE',
                                                 'ITEM DESCRIPTION',
                                                 'PRICE',
                                                 'SALES TAX',
                                                 'VENDOR PAYOUT',
                                                 'BEV PAYOUT',
                                                 'DATE',
                                                 'TIME']]
    item_transactions_df['DATE'] = [datetime.strptime(date_val, "%Y-%m-%d") for date_val in item_transactions_df['DATE']]
    item_transactions_df['TIME'] = [datetime.strptime(time_val, "%H:%M:%S") for time_val in item_transactions_df['TIME']]
    wb = load_workbook(filename = f'The Beverly Collective Sales {datetime.now().year}.xlsx')
    ws = wb.create_sheet(new_sheetname, -1) # adds sheet as second-to-last sheet in workbook (last is vendor info)
    for r in dataframe_to_rows(item_transactions_df, index=False, header=True):
        ws.append(r)
    for row in range(1, len(item_transactions_df)+2):
        ws[f'I{row}'].number_format='mm-dd-yy'
        ws[f'J{row}'].number_format='h:mm:ss'
        ws[f'E{row}'].number_format='"$"#,##0.00'
        ws[f'F{row}'].number_format='"$"#,##0.00'
        ws[f'G{row}'].number_format='"$"#,##0.00'
        ws[f'H{row}'].number_format='"$"#,##0.00'
        if ws[f'B{row}'].value=='NONE':
            ws[f'B{row}'].fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        for col in ['G', 'H']:
            if ws[f'{col}{row}'].value=='Error':
                ws[f'{col}{row}'].fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        if ws[f'D{row}'].value=='':
            ws[f'D{row}'].fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        ws[f'{col}1'].font = Font(bold=True)
    wb.save(f"C:/Users/Ryan/BevCol/The Beverly Collective Sales {datetime.now().year}.xlsx")
    return item_transactions_df


def main():
    os.chdir('C:/Users/Ryan/BevCol/')
    (image_date,
     new_sheetname,
     period_start_date,
     period_end_date,
     last_full_report,
     vendor_info_df,
     all_vendors_list,
     square_download_filename) = gather_input_info()
    item_transactions_df = preprocess_item_details_square_file(square_download_filename=square_download_filename,
                                                               first_transaction_num=(max(last_full_report['TRANSACTION'])+1),
                                                               vendor_info_df=vendor_info_df,
                                                               new_sheetname=new_sheetname)
    _ = input(f'Press enter after reviewing The Beverly Collective Sales {datetime.now().year} file and cleaning missing values.')
    current_full_report = pd.read_excel(f'C:/Users/Ryan/BevCol/The Beverly Collective Sales {datetime.now().year}.xlsx',
                                        new_sheetname)
    grouped_df = calculate_summary_amount(current_full_report)
    create_image_attachments(current_full_report,
                             image_date)
    service = establish_gmail_api_connection()
    create_vendor_email_drafts(service,
                               current_full_report,
                               grouped_df,
                               vendor_info_df,
                               period_start_date,
                               period_end_date,
                               image_date,
                               all_vendors_list)
    # deleting to avoid linting suggestions, but returned the
    # item_transactions_df to enable troubleshooting if needed
    del item_transactions_df
    return


if __name__=='__main__':
    main()
